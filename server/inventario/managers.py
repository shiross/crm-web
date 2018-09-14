import hashlib
import uuid
from datetime import timedelta, datetime
from time import time
from sqlalchemy.exc import IntegrityError

from server.common.managers import SuperManager
from .models import *
from ..user.models import User
from openpyxl import load_workbook, Workbook


class CompanyManager(SuperManager):
    def __init__(self, db):
        super().__init__(Company, db)

    def getCompanyByID(self, id):
        return self.db.query(Company).filter(Company.id == id).first()

    def list_all(self):
        return self.db.query(Company).filter(Company.id > 0)

    def import_excel(self, cname):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            sheet = wb.get_sheet_by_name(name='Hoja1')  # parametrizable......
            colnames = ['IdEmpresa', 'NomEmpresa', 'NitEmpresa', 'ActEmpresa',
                        'PagEmpresa', 'TxtEmpresa', 'TcEmpresa']
            min_row = 1
            indices = {cell[0].value: n - min_row for n, cell in
                       enumerate(sheet.iter_cols(min_row=min_row, max_row=min_row), start=min_row) if
                       cell[0].value in colnames}
            for row in sheet.iter_rows(min_row=min_row + 1):
                if row[indices['IdEmpresa']].value is not None and \
                                row[indices['NomEmpresa']].value is not None and \
                                row[indices['NitEmpresa']].value is not None and \
                                row[indices['ActEmpresa']].value is not None and \
                                row[indices['PagEmpresa']].value is not None and \
                                row[indices['TxtEmpresa']].value is not None and \
                                row[indices['TcEmpresa']].value is not None:

                    if CompanyManager(self.db).getCompanyByID(row[indices['IdEmpresa']].value) is None:
                        empresa = Company(
                            id=row[indices['IdEmpresa']].value,
                            name=row[indices['NomEmpresa']].value,
                            nit=row[indices['NitEmpresa']].value,
                            activity=row[indices['ActEmpresa']].value,
                            payment=row[indices['PagEmpresa']].value,
                            text=row[indices['PagEmpresa']].value,
                            coin=row[indices['TcEmpresa']].value
                        )
                        self.db.add(empresa)
                        self.db.flush()
            self.db.commit()
            return {'message': 'Importado Todos Correctamente.', 'success': True}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint failed: rrhh_persona.dni' in str(e):
                return {'message': 'CI duplicado', 'success': False}
            if 'UNIQUE constraint failed: rrhh_empleado.codigo' in str(e):
                return {'message': 'codigo de empleado duplicado', 'success': False}
            return {'message': str(e), 'success': False}


class OfficeManager(SuperManager):
    def __init__(self, db):
        super().__init__(Office, db)

    def getOfficeByID(self, id):
        return self.db.query(Office).filter(Office.id == id).first()

    def getOfficeByName(self, name):
        return self.db.query(Office).filter(Office.name == name).first()

    def getOfficeByPromoter(self, fk_promoter):
        return self.db.query(Office).filter(Office.fk_promoter == fk_promoter).first()

    def listAll(self):
        return self.db.query(Office).filter(Office.id > 0)

    def import_excel(self, cname):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            sheet = wb.get_sheet_by_name(name='Hoja1')  # parametrizable......
            colnames = ['idEmpresa', 'idSucursal', 'NomSucursal', 'DirSucusal',
                        'TelSucursal', 'DepSucursal', 'LdoSucursal',
                        'NauSucursal', 'DoiSucursal', 'DofSucursal', 'FliSucursal', 'idPromotor']
            min_row = 1
            indices = {cell[0].value: n - min_row for n, cell in
                       enumerate(sheet.iter_cols(min_row=min_row, max_row=min_row), start=min_row) if
                       cell[0].value in colnames}
            for row in sheet.iter_rows(min_row=min_row + 1):
                if row[indices['idEmpresa']].value is not None and \
                                row[indices['idSucursal']].value is not None and \
                                row[indices['NomSucursal']].value is not None and \
                                row[indices['DirSucusal']].value is not None and \
                                row[indices['TelSucursal']].value is not None and \
                                row[indices['DepSucursal']].value is not None and \
                                row[indices['LdoSucursal']].value is not None and \
                                row[indices['NauSucursal']].value is not None and \
                                row[indices['DoiSucursal']].value is not None and \
                                row[indices['DofSucursal']].value is not None and \
                                row[indices['FliSucursal']].value is not None and \
                                row[indices['idPromotor']].value is not None:

                    if OfficeManager(self.db).getOfficeByID(row[indices['idSucursal']].value) is None:
                        sucursal = Office(
                            id=row[indices['idSucursal']].value,
                            name=row[indices['NomSucursal']].value,
                            address=row[indices['DirSucusal']].value,
                            phone=row[indices['TelSucursal']].value,
                            department=row[indices['DepSucursal']].value,
                            key=row[indices['LdoSucursal']].value,
                            auth=row[indices['NauSucursal']].value,
                            dsfini=row[indices['DoiSucursal']].value,
                            dsffin=row[indices['DofSucursal']].value,
                            limdate=row[indices['FliSucursal']].value,
                            fk_company=row[indices['idEmpresa']].value,
                            fk_promoter=row[indices['idPromotor']].value
                        )
                        self.db.add(sucursal)
                        self.db.flush()
            self.db.commit()
            return {'message': 'Importado Todos Correctamente.', 'success': True}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint failed: rrhh_persona.dni' in str(e):
                return {'message': 'CI duplicado', 'success': False}
            if 'UNIQUE constraint failed: rrhh_empleado.codigo' in str(e):
                return {'message': 'codigo de empleado duplicado', 'success': False}
            return {'message': str(e), 'success': False}


class ProductManager(SuperManager):
    def __init__(self, db):
        super().__init__(Product, db)

    # def get_groupsByFamily(self, family):
    #     return self.db.query(Product).filter(Product.fk_family == family)

    def getProductByID(self, id):
        return self.db.query(Product).filter(Product.id == id).first()

    def getProductByName(self, name):
        return self.db.query(Product).filter(Product.name == name).first()

    def listAll(self):
        return self.db.query(Product).filter(Product.id > 0)

    def import_excel(self, cname):
        try:
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            sheet = wb.get_sheet_by_name(name='Hoja1')  # parametrizable......
            colnames = ['idEmpresa', 'idProducto', 'NomProducto', 'uniProducto',
                        'PreProducto', 'desProducto', 'salProducto']
            min_row = 1
            indices = {cell[0].value: n - min_row for n, cell in
                       enumerate(sheet.iter_cols(min_row=min_row, max_row=min_row), start=min_row) if
                       cell[0].value in colnames}
            for row in sheet.iter_rows(min_row=min_row + 1):
                if row[indices['idEmpresa']].value is not None and \
                                row[indices['idProducto']].value is not None and \
                                row[indices['NomProducto']].value is not None and \
                                row[indices['uniProducto']].value is not None and \
                                row[indices['PreProducto']].value is not None and \
                                row[indices['desProducto']].value is not None and \
                                row[indices['salProducto']].value is not None:

                    if ProductManager(self.db).getProductByID(row[indices['idProducto']].value) is None:
                        producto = Product(
                            id=row[indices['idProducto']].value,
                            name=row[indices['NomProducto']].value,
                            unit=row[indices['uniProducto']].value,
                            price=row[indices['PreProducto']].value,
                            discount=row[indices['desProducto']].value,
                            stock=row[indices['salProducto']].value,
                            fk_company=row[indices['idEmpresa']].value
                        )
                        self.db.add(producto)
                        self.db.flush()
            self.db.commit()
            return {'message': 'Importado Todos Correctamente.', 'success': True}

        except IntegrityError as e:
            self.db.rollback()
        if 'UNIQUE constraint failed: rrhh_persona.dni' in str(e):
            return {'message': 'CI duplicado', 'success': False}
        if 'UNIQUE constraint failed: rrhh_empleado.codigo' in str(e):
            return {'message': 'codigo de empleado duplicado', 'success': False}
        return {'message': str(e), 'success': False}


class CustomerManager(SuperManager):
    def __init__(self, db):
        super().__init__(Customer, db)

    def getCustomerByID(self, id):
        return self.db.query(Customer).filter(Customer.id == id).first()

    def listAll(self):
        return self.db.query(Customer).filter(Customer.id > 0)

    def getCustomerByName(self, line):
        return self.db.query(Customer).filter(Customer.name == line).first()

    def getCustomerByNit(self, family):
        return self.db.query(Customer).filter(Customer.nit == family)

    def import_excel(self, cname):
        try:
            print("init flaco")
            print(str(datetime.now()))
            wb = load_workbook(filename="server/common/resources/uploads/" + cname)
            sheet = wb.get_sheet_by_name(name='Hoja1')  # parametrizable......
            colnames = ['idEmpresa', 'idCliente', 'NomCliente', 'NitCliente']
            min_row = 1
            indices = {cell[0].value: n - min_row for n, cell in
                       enumerate(sheet.iter_cols(min_row=min_row, max_row=min_row), start=min_row) if
                       cell[0].value in colnames}
            # print(str(datetime.now()))
            for row in sheet.iter_rows(min_row=min_row + 1):
                if row[indices['idEmpresa']].value is not None and \
                                row[indices['idCliente']].value is not None and \
                                row[indices['NomCliente']].value is not None and \
                                row[indices['NitCliente']].value is not None:

                    if CustomerManager(self.db).getCustomerByID(row[indices['idCliente']].value) is None:
                        cliente = Customer(
                            id=row[indices['idCliente']].value,
                            name=row[indices['NomCliente']].value,
                            nit=row[indices['NitCliente']].value,
                            fk_company=row[indices['idEmpresa']].value
                        )
                        self.db.add(cliente)
                        self.db.flush()

            self.db.commit()
            print("end flaco")
            print(str(datetime.now()))
            return {'message': 'Importado Todos Correctamente.', 'success': True}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint failed: rrhh_persona.dni' in str(e):
                return {'message': 'CI duplicado', 'success': False}
            if 'UNIQUE constraint failed: rrhh_empleado.codigo' in str(e):
                return {'message': 'codigo de empleado duplicado', 'success': False}
            return {'message': str(e), 'success': False}


class BillManager(SuperManager):
    def __init__(self, db):
        super().__init__(Bill, db)

    def getBillByID(self, id):
        return self.db.query(Bill).filter(Bill.id == id).first()

    # def get_subgroupByLine(self, line):
    #     return self.db.query(SubGroup).filter(SubGroup.subline == line).first()
    #
    # def get_subgroupsByGroup(self, group):
    #     return self.db.query(SubGroup).filter(SubGroup.fk_group == group)

    def insertBills(self, data):
        #     data es una lista de bills :v
        try:
            for bill in data:

                managedBill = Bill(ctrlcode=bill['code'],
                                   id_bill=bill['idBill'],
                                   fk_company=bill['fkCompany'],
                                   fk_customer=bill['fkCustomer'],
                                   date=datetime.strptime(bill['date'], '%Y/%m/%d').date(),
                                   hour=datetime.strptime(bill['hour'], '%H:%M:%S').time(),
                                   nit=bill['nit'],
                                   name=bill['name'],
                                   fk_promoter=bill['fkPromoter']
                                   )
                for detail in bill['detailList']:
                    managedDetail = Detail(fk_bill=detail['fkBill'],
                                           fk_product=detail['fkProduct'],
                                           quantity=detail['quantity'],
                                           oriprize=detail['iprice'],
                                           finprize=detail['fprice'],
                                           total=detail['total'])
                    managedBill.details.append(managedDetail)
                self.db.add(managedBill)
            self.db.commit()
            return True
        except IntegrityError as e:
            print(str(e))
            self.db.rollback()
            return False

    def listAll(self):
        return self.db.query(Bill).filter(Bill.id > 0)


class DetailManager(SuperManager):
    def __init__(self, db):
        super().__init__(Detail, db)

    def getDetailByID(self, code):
        return self.db.query(Detail).filter(Detail.id == code).first()

    def listAll(self):
        return self.db.query(Detail).filter(Detail.id > 0)

        # def get_productsBySubgroup(self, subgroup):
        #     return self.db.query(Detail).filter(Detail.fk_subgroup == subgroup)
        #
        # def get_productsByGroup(self, group):
        #     return self.db.query(Detail).filter(Detail.fk_group == group.id_group).filter(
        #         Detail.fk_family == group.fk_family)
        #
        # def get_productsByFamily(self, family):
        #     return self.db.query(Detail).filter(Detail.fk_family == family)
        #
        # def get_productsByWarehouse(self, warehouse):
        #     return self.db.query(Detail).filter(Detail.fk_warehouse == warehouse)
        #
        # def get_productByCode(self, code):
        #     return self.db.query(Detail).filter(Detail.code == code).first()
        #
        # def get_productByOrigin(self, origin):
        #     return self.db.query(Detail).filter(Detail.code_origin == origin)
        #
        # def get_productsByText(self, text):
        #     return self.db.query(Detail).filter(self.colums_like(self.entity, text))

        # def import_excel(self, cname):
        #     try:
        #         wb = load_workbook(filename="server/common/resources/uploads/" + cname)
        #         sheet = wb.get_sheet_by_name(name='Hoja1')  # parametrizable......
        #         colnames = ['almacen', 'nomb_alma', 'id_fami', 'nomb_fami', 'id_grup',
        #                     'line', 'nomb_grup', 'id_subg', 'subline', 'nomb_subg', 'codigo',
        #                     'codorigen', 'nomb_cata', 'fechacaduc', 'nrolote', 'nrofabri', 'saldf_ini', 'impor_ini',
        #                     'ingresos', 'salidas',
        #                     'saldf_fin']
        #         min_row = 1
        #         indices = {cell[0].value: n - min_row for n, cell in
        #                    enumerate(sheet.iter_cols(min_row=min_row, max_row=min_row), start=min_row) if
        #                    cell[0].value in colnames}
        #         for row in sheet.iter_rows(min_row=min_row + 1):
        #             if row[indices['almacen']].value is not None and row[indices['nomb_alma']].value is not None and \
        #                             row[indices['id_fami']].value is not None and \
        #                             row[indices['nomb_fami']].value is not None and \
        #                             row[indices['id_grup']].value is not None and \
        #                             row[indices['line']].value is not None and \
        #                             row[indices['nomb_grup']].value is not None and \
        #                             row[indices['id_subg']].value is not None and \
        #                             row[indices['subline']].value is not None and \
        #                             row[indices['nomb_subg']].value is not None and \
        #                             row[indices['codigo']].value is not None and \
        #                             row[indices['codorigen']].value is not None and \
        #                             row[indices['saldf_fin']].value is not None and \
        #                             row[indices['nomb_cata']].value is not None:
        #
        #                 if WarehouseManager(self.db).get_warehouseByID(row[indices['almacen']].value) is None:
        #                     almacen = Warehouse(
        #                         id=row[indices['almacen']].value,
        #                         name=row[indices['nomb_alma']].value
        #                     )
        #
        #                 if FamilyManager(self.db).get_familyByID(row[indices['id_fami']].value) is None:
        #                     familia = Family(
        #                         id=row[indices['id_fami']].value,
        #                         name=row[indices['nomb_fami']].value
        #                     )
        #
        #                 if GroupManager(self.db).get_groupByLine(row[indices['line']].value) is None:
        #                     grupo = Group(
        #                         id_group=row[indices['id_grup']].value,
        #                         line=row[indices['line']].value,
        #                         name=row[indices['nomb_grup']].value
        #                     )
        #                 if SubGroupManager(self.db).get_subgroupByLine(row[indices['subline']].value) is None:
        #                     subgrupo = SubGroup(
        #                         id_subgroup=row[indices['id_subg']].value,
        #                         subline=row[indices['subline']].value,
        #                         name=row[indices['nomb_subg']].value
        #                     )
        #                 if ProductManager(self.db).get_product(row[indices['codigo']].value) is None:
        #                     producto = InvProduct(
        #                         code=row[indices['codigo']].value,
        #                         code_origin=row[indices['codorigen']].value,
        #                         name=row[indices['nomb_cata']].value,
        #                         stock=row[indices['saldf_fin']].value
        #                     )
        #                 else:
        #                     producto = ProductManager(self.db).get_product(row[indices['codigo']].value)
        #                     producto.stock = producto.stock + row[indices['saldf_fin']].value
        #
        #                 if FamilyGroupManager(self.db).get_familygroup(familia.id, grupo.line) is None:
        #                     familiagrupo = FamilyGroup()
        #                 else:
        #                     familiagrupo = FamilyGroupManager(self.db).get_familygroup(familia.id, grupo.line)
        #
        #                 if GroupSubgroupManager(self.db).get_groupsub(grupo.line, subgrupo.subline) is None:
        #                     gruposubgrupo = GroupSubGroup()
        #                 else:
        #                     gruposubgrupo = GroupSubgroupManager(self.db).get_groupsub(grupo.line, subgrupo.subline)
        #                 producto.subgroup = subgrupo
        #                 # producto.group = grupo
        #                 # producto.family = familia
        #                 # producto.warehouse = almacen
        #                 # subgrupo.group = grupo
        #                 # grupo.family = familia
        #                 familiagrupo.family = familia
        #                 familiagrupo.group = grupo
        #                 gruposubgrupo.group = grupo
        #                 gruposubgrupo.subgroup = subgrupo
        #                 familia.warehouse = almacen
        #
        #                 self.db.add(producto)
        #                 self.db.add(familiagrupo)
        #                 self.db.add(gruposubgrupo)
        #                 self.db.flush()
        #
        #             self.db.commit()
        #         return {'message': 'Importado Todos Correctamente.', 'success': True}
        #     except IntegrityError as e:
        #         self.db.rollback()
        #         if 'UNIQUE constraint failed: rrhh_persona.dni' in str(e):
        #             return {'message': 'CI duplicado', 'success': False}
        #         if 'UNIQUE constraint failed: rrhh_empleado.codigo' in str(e):
        #             return {'message': 'codigo de empleado duplicado', 'success': False}
        #         return {'message': str(e), 'success': False}
